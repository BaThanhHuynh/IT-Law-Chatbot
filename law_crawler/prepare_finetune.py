"""
prepare_finetune.py
===================
Chuẩn bị corpus sạch cho Thành fine-tune PhoBERT MLM.
Ưu tiên full_dieu_text từ law_chunks_hier.jsonl + Nội dung điều từ Excel.

Output: data/finetune_corpus.jsonl  →  {"text": "toàn bộ điều luật dài"}
"""

import json
from pathlib import Path
import openpyxl
from collections import defaultdict

# ========================== CẤU HÌNH ==========================
EXCEL_PATH     = "data/law_data_output.xlsx"
HIER_JSONL_PATH = "data/law_chunks_hier.jsonl"   
OUTPUT_PATH    = "data/finetune_corpus.jsonl"
MIN_LEN        = 50
# ============================================================

def main():
    corpus = []
    seen = set()          # dedup theo 200 ký tự đầu

    # ==================== 1. LẤY FULL_DIEU_TEXT TỪ HIER ====================
    print(f"📖 Đọc full_dieu_text từ {HIER_JSONL_PATH}...")
    full_dieu_count = 0
    with open(HIER_JSONL_PATH, "r", encoding="utf-8") as f:
        for line in f:
            item = json.loads(line.strip())
            full_text = item.get("payload", {}).get("full_dieu_text", "").strip()
            if len(full_text) >= MIN_LEN:
                key = full_text[:200]  # dedup
                if key not in seen:
                    seen.add(key)
                    corpus.append(full_text)
                    full_dieu_count += 1
    print(f"  → {full_dieu_count:,} full điều luật từ hierarchical JSONL")

    # ==================== 2. BỔ SUNG TỪ EXCEL (phòng trường hợp thiếu) ====================
    print(f"📖 Đọc Nội dung điều từ {EXCEL_PATH}...")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb["Dữ liệu luật"]
    headers = [str(h).strip() if h else "" for h in next(ws.iter_rows(max_row=1, values_only=True))]
    hi = {h: i for i, h in enumerate(headers)}

    excel_count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        nd = str(row[hi.get("Nội dung điều", 0)] or "").strip()
        if len(nd) >= MIN_LEN:
            key = nd[:200]
            if key not in seen:
                seen.add(key)
                corpus.append(nd)
                excel_count += 1
    print(f"  → {excel_count:,} điều bổ sung từ Excel")

    # ==================== 3. THỐNG KÊ & XUẤT FILE ====================
    print(f"\n✅ Tổng corpus sau dedup: {len(corpus):,} mẫu")
    total_chars = sum(len(t) for t in corpus)
    est_tokens = total_chars // 2
    print(f"   Tổng ký tự: {total_chars:,}")
    print(f"   Ước tính tokens: ~{est_tokens/1_000_000:.1f}M tokens")

    Path(OUTPUT_PATH).parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        for text in corpus:
            f.write(json.dumps({"text": text}, ensure_ascii=False) + "\n")

    print(f"\n🎉 ĐÃ XUẤT: {OUTPUT_PATH}")
    print("   → Gửi file này cho Thành fine-tune PhoBERT MLM là chuẩn nhất!")

if __name__ == "__main__":
    main()